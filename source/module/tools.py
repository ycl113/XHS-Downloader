from asyncio import sleep
from random import uniform

from rich import print
from rich.text import Text

from .static import INFO
import pandas as pd
from openpyxl import load_workbook


def retry(function):
    async def inner(self, *args, **kwargs):
        if result := await function(self, *args, **kwargs):
            return result
        for _ in range(self.retry):
            if result := await function(self, *args, **kwargs):
                return result
        return result

    return inner


def logging(log, text, style=INFO):
    string = Text(text, style=style)
    if log:
        log.write(string, animate=True, scroll_end=True, )
    else:
        print(string)


async def sleep_time(
        min_time: int | float = 0.5,
        max_time: int | float = 1.5,
):
    await sleep(uniform(min_time, max_time))


async def save_xls(data, path):
    df = pd.DataFrame([data])

    # 定义 Excel 文件名
    excel_file = path + '/xhs_data.xlsx'

    # 追加数据到 Excel 文件
    try:
    # 尝试加载现有工作簿
        book = load_workbook(excel_file)
        with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            # 获取现有工作表的最大行数
            start_row = writer.sheets['Sheet1'].max_row
            # 写入数据
            df.to_excel(writer, index=False, header=False, startrow=start_row)
            return excel_file
    except FileNotFoundError:
        # 如果文件不存在，创建一个新的文件
        df.to_excel(excel_file, index=False)
